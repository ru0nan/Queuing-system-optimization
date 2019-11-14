# -*- coding: utf-8 -*-
"""
Created on Sat Mar 10 09:08:46 2018

@author: lenovo
"""

import numpy as np
import random
import sys
import time
from openpyxl import load_workbook

#    各优先级对应的wait time target（day）
dic={1:1,2:2,3:10,4:28}
pi ={1:28/41,2:10/41,3:2/41,4:1/41}



class Patient:
    alpha=1.2598731
    beta=1.84471708
    def __init__(self,pri,arr_date):
        self.init_pri = pri
        self.acc_wei = pri
        self.arr_date = arr_date
        self.WTT = dic[pri]
        self.wait_time = 0
        self.exceedFlag = 0
        
    def update_weight(self):
#        self.acc_wei = alpha*(self.init_pri) + beta*self.wait_time
        self.acc_wei = Patient.alpha*pi[self.init_pri] +\
        Patient.beta*self.wait_time
        
    def update_wait_time(self):   
        self.wait_time += 1
#        if not(self.exceedFlag) and self.wait_time > self.WTT:
#            self.exceedFlag = 1

def arr_pre():
    num_total = int(random.normalvariate(41.3,12.8))
#    print('this is today\'s arrival: ',end='')
#    print(arrive)
    return num_total

def ser_pre():
#    正态分布
    served = int(random.normalvariate(30.05,7.47))
#    weibull分布 scale param(λ)=32.89,shap param(k)=4.58
#    served = int(random.weibullvariate(32.89,4.58))
    return served
   
#更新“优先级-等待天数”矩阵
def add_up(patient):
    global pri_wt
    pri_wt[patient.wait_time][patient.init_pri] += 1
   

def que_end(que,duration):
    WTinQ = np.zeros((duration+1,5),dtype=int)
    for i in range(len(que)):
        WTinQ[que[i].wait_time][que[i].init_pri] += 1
    return WTinQ

def sum_exc_time(pw_mat):
    total_ET = 0
    for i in [1,2,3,4]:
        for j in range(len(pw_mat)):
            if pw_mat[j][i]>0 and j > dic[i]:
                total_ET += (j-dic[i])*pw_mat[j][i]
    return total_ET

def overflow(pw_mat,pri_arr_num):
#    pro = {1:0,2:0,3:0,4:0} #各个优先级对应的overflow proportion
    pro = [0,0,0,0,0]
    exc = [0,0,0,0,0]#各个优先级患者超出WTT的人数
    for i in [1,2,3,4]:
        exc[i] = sum(pw_mat[dic[i]+1:pw_mat.shape[0],i])
    for i in range(1,5):
        pro[i] = exc[i]/pri_arr_num[i]
    return pro,exc
        
 
def view_bar(num, total):
    rate = num / total
    rate_num = int(rate * 100)
    r = '\r[%s%s]%d%%' % ("="*int(num/total*10), " "*(10-num), rate_num, )
    sys.stdout.write(r)
    sys.stdout.flush()   

def output(a=6,b=1):

    readFile = 'D:\\学习\\毕设\\排队模型python\\WTIS统计.xlsx'

    wb = load_workbook(readFile,read_only = True)
    ws = wb.get_sheet_by_name('Sheet1')
    sheet = wb.active
    
    ws_rows_len = sheet.max_row   #len(list(ws.rows))          #行数
    ws_col_len = sheet.max_column #len(list(ws.columns))    #列数
    temp_data=[[0 for i in range(ws_col_len)] for j in range(ws_rows_len-1)]
    
    for i in range(2, ws_rows_len):
        for j in range(1, ws_col_len+1):
            temp_data[i-2][j-1] = ws.cell(row=i,column=j).value
    
    Patient.alpha = a
    Patient.beta = b
    queue=[]
    duration = len(temp_data) #仿真天数
    pri_arr_num = [0,0,0,0,0] #统计仿真期间各优先级的到达情况
    global pri_wt
    pri_wt = np.zeros((duration+1,5),dtype=int)#优先级-等待时间 矩阵

    for i in range(duration):
        date = i+1
        #每日患者的到达情况，arrive是优先级-人数字典
        num_p1 = temp_data[i][0]
        num_p2 = temp_data[i][1]
        num_p3 = temp_data[i][2]
        num_p4 = temp_data[i][3]
        arrive = {1:num_p1,2:num_p2,3:num_p3,4:num_p4}

        pri_arr_num[1] += arrive[1]
        pri_arr_num[2] += arrive[2]
        pri_arr_num[3] += arrive[3]
        pri_arr_num[4] += arrive[4]
        #将新来的患者入队
        for j in [1,2,3,4]:
            for k in range(arrive[j]):
                queue.append(Patient(j,date))

        #将新来的患者入队后，给队列每个人更新权重
        for j in range(len(queue)):
            queue[j].update_weight()

        queue = sorted(queue, key=lambda patient: patient.acc_wei,reverse = True)  # sort by weight
        
        #出队并统计等待情况                
        #num_ser = ser_pre()
        num_ser = temp_data[i][4]
        
        for j in range(num_ser):
            if len(queue)>0:
                add_up(queue[0])
                del queue[0]
            else:
#                print('queue is empty')
                break
            
        for j in range(len(queue)):
            queue[j].update_wait_time()
            
    WTinQ = que_end(queue,duration)
    pw_mat = WTinQ+pri_wt #结束时队列患者的等待时间+出队患者的等待时间
    total_ET = sum_exc_time(pw_mat)
    (ofp,exc) = overflow(pw_mat,pri_arr_num)
#    ofp1 = (ofp[1]*28+ofp[2]*10+ofp[3]*2+ofp[4])/41
#    ser_pro =  sum(sum(pri_wt))/sum(pri_arr_num)
#    return (total_ET,ofp1,ser_pro)
    return total_ET

  
#    内部测试
if __name__ == "__main__":
#    time_start=time.time();#time.time()为1970.1.1到当前时间的毫秒数  
#    queue=[]
#    duration = 300  #仿真天数
#    pri_arr_num = [0,0,0,0,0] #统计仿真期间各优先级的到达情况
#    global pri_wt
#    pri_wt = np.zeros((duration+1,5),dtype=int)#优先级-等待时间 矩阵
#    
##    pri_ser_num =  {1:0,2:0,3:0,4:0}  #字典，已服务患者优先级-人数
##    exc_time = {1:0,2:0,3:0,4:0}  #优先级-超出WTT的时间
##    exc_num  = {1:0,2:0,3:0,4:0}  #优先级-超出WTT的人数
#
#    for i in range(duration):
#        date = i+1
##        print('date',date)
#        arrive = arr_pre()#每日患者的到达情况，arrive是优先级-人数字典
#        pri_arr_num[1] += arrive[1]
#        pri_arr_num[2] += arrive[2]
#        pri_arr_num[3] += arrive[3]
#        pri_arr_num[4] += arrive[4]
#        #将新来的患者入队
#        for j in [1,2,3,4]:
#            for k in range(arrive[j]):
#                queue.append(Patient(j,date))
#
#        #将新来的患者入队后，给队列每个人更新权重
#        for j in range(len(queue)):
#            queue[j].update_weight()
#
## =============================================================================
##        print('before sorting:',end='')
##        for j in range(len(queue)):
##            print('[{0:2d} {1:2d}]'.format(queue[j].init_pri,queue[j].wait_time),end=' ')
##        print('')
## =============================================================================
#        queue = sorted(queue, key=lambda patient: patient.acc_wei,reverse = True)  # sort by weight
#        
## =============================================================================
##        print('after sorting:',end='')
##        for j in range(len(queue)):
##            print('[{0:2d} {1:2d}]'.format(queue[j].init_pri,queue[j].wait_time),end=' ')
##        print('')
## =============================================================================
#        #出队并统计等待情况                
#        num_ser = ser_pre()
#        for j in range(num_ser):
#            if len(queue)>0:
#                add_up(queue[0])
#                del queue[0]
#            else:
#                print('queue is empty')
#                break
#            
#        for j in range(len(queue)):
#            queue[j].update_wait_time()
#            
#        view_bar(i+1, duration)
#
#    
#    WTinQ = que_end(queue,duration)
#    pw_mat = WTinQ+pri_wt #结束时队列患者的等待时间+出队患者的等待时间
#    total_ET = sum_exc_time(pw_mat)
#    (ofp,exc) = overflow(pw_mat,pri_arr_num)
#    total_num_exc = sum(exc)
#    ofp1 = (ofp[1]*28+ofp[2]*10+ofp[3]*2+ofp[4])/41
#    ser_pro =  sum(sum(pri_wt))/sum(pri_arr_num)
#    print('\nend')
#
#  
#    time_end=time.time();#time.time()为1970.1.1到当前时间的秒数  
#    print(time_end-time_start,end='')
#    print('s')
    
    readFile = 'D:\\学习\\毕设\\排队模型python\\WTIS统计.xlsx'

    wb = load_workbook(readFile,read_only = True)
    ws = wb.get_sheet_by_name('Sheet1')
    sheet = wb.active
    
    ws_rows_len = sheet.max_row   #len(list(ws.rows))          #行数
    ws_col_len = sheet.max_column #len(list(ws.columns))    #列数
    temp_data=[[0 for i in range(ws_col_len)] for j in range(ws_rows_len-1)]
    
    for i in range(2, ws_rows_len):
        for j in range(1, ws_col_len+1):
            temp_data[i-2][j-1] = ws.cell(row=i,column=j).value

    
    Patient.alpha = 6
    Patient.beta = 1
    queue=[]
    duration = len(temp_data) #仿真天数
    pri_arr_num = [0,0,0,0,0] #统计仿真期间各优先级的到达情况
    global pri_wt
    pri_wt = np.zeros((duration+1,5),dtype=int)#优先级-等待时间 矩阵

    for i in range(duration):
        date = i+1
        #每日患者的到达情况，arrive是优先级-人数字典
        #num_total = arr_pre()
        
        
        num_p1 = temp_data[i][0]
        num_p2 = temp_data[i][1]
        num_p3 = temp_data[i][2]
        num_p4 = temp_data[i][3]
        arrive = {1:num_p1,2:num_p2,3:num_p3,4:num_p4}

        pri_arr_num[1] += arrive[1]
        pri_arr_num[2] += arrive[2]
        pri_arr_num[3] += arrive[3]
        pri_arr_num[4] += arrive[4]
        #将新来的患者入队
        for j in [1,2,3,4]:
            for k in range(arrive[j]):
                queue.append(Patient(j,date))

        #将新来的患者入队后，给队列每个人更新权重
        for j in range(len(queue)):
            queue[j].update_weight()

        queue = sorted(queue, key=lambda patient: patient.acc_wei,reverse = True)  # sort by weight
        
        #出队并统计等待情况                
        #num_ser = ser_pre()
        num_ser = temp_data[i][4]
        
        for j in range(num_ser):
            if len(queue)>0:
                add_up(queue[0])
                del queue[0]
            else:
#                print('queue is empty')
                break
            
        for j in range(len(queue)):
            queue[j].update_wait_time()
            
    WTinQ = que_end(queue,duration)
    pw_mat = WTinQ+pri_wt #结束时队列患者的等待时间+出队患者的等待时间
    total_ET = sum_exc_time(pw_mat)
    (ofp,exc) = overflow(pw_mat,pri_arr_num)
    print('end')
    
# =============================================================================
#        逐行读取大文件        
# i = 1
# with open('D:\\学习\\毕设\\数据资料\\MRI_Booking2014.txt', 'r', encoding = 'utf-8') as f: 
#     for line in f:
#         print(line)
#         i+=1
#         if i>4:
#             break
# =============================================================================


# 带有异常捕捉的更新统计量函数
#def add_up(patient):
#    try:
#        global sum_wt, sum_exceed, sum_served, exc_time, exc_num
#        sum_served += 1
#        sum_wt += patient.wait_time
#        if patient.exceedFlag:
#            try:
#                sum_exceed += patient.exceedFlag
#                exc_time[patient.init_pri] += patient.wait_time - patient.WTT
#                exc_num[patient.init_pri] += 1
#            except:
#                print('error in if:',sys.exc_info()[0])
#            else:
#                pass
#    except:
#        print('error in mainbody:',sys.exc_info()[0])
#    else:
#        pass
